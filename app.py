import os
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, flash
import psycopg2
import psycopg2.extras
from datetime import datetime
import io
import openpyxl
from psycopg2.extras import RealDictCursor
import pdfkit
import shutil
from urllib.parse import urlparse

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your_secret_key_here')  # Use environment variable for production

# Database connection setup
def get_db_connection():
    # Parse DATABASE_URL for Render deployment
    database_url = os.environ.get('DATABASE_URL')
    
    if database_url:
        # Handle Render's PostgreSQL URL format
        if database_url.startswith("postgres://"):
            database_url = database_url.replace("postgres://", "postgresql://", 1)
        
        # Parse the database URL
        result = urlparse(database_url)
        conn = psycopg2.connect(
            dbname=result.path[1:],
            user=result.username,
            password=result.password,
            host=result.hostname,
            port=result.port,
            sslmode='require'  # Important for Render PostgreSQL
        )
        return conn
    else:
        # Fallback to local config (for development)
        return psycopg2.connect(
            host='localhost',
            dbname='postgres',
            user='postgres',
            password='12Marks@255'
        )

# Set up pdfkit configuration
WKHTMLTOPDF_PATH = shutil.which("wkhtmltopdf")
if WKHTMLTOPDF_PATH is None:
    WKHTMLTOPDF_PATH = r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"  # for local use

pdfkit_config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

# Helper for dict-like row access
def fetchall_dict(cur):
    return cur.fetchall()

def fetchone_dict(cur):
    return cur.fetchone()

# Home redirects to login
@app.route('/')
def home():
    return redirect(url_for('login'))

# Login Page
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        try:
            cur.execute("SELECT * FROM users WHERE username=%s AND password=%s", (username, password))
            user = cur.fetchone()
        except Exception as e:
            error = f"Database error: {e}"
            user = None
        conn.close()
        if user:
            session['user'] = username
            session['role'] = user.get('role', 'user')
            return redirect(url_for('dashboard'))
        else:
            if not error:
                error = 'Invalid Username or Password'
    return render_template('login.html', error=error)

# Registration Page (add role selection, default to 'user')
@app.route('/register', methods=['GET', 'POST'])
def register():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form.get('role', 'user')
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM users WHERE username=%s", (username,))
        user = cur.fetchone()
        if user:
            error = "Username already exists."
        else:
            cur = conn.cursor()
            cur.execute("INSERT INTO users (username, password, role) VALUES (%s, %s, %s)", (username, password, role))
            conn.commit()
            conn.close()
            return redirect(url_for('login'))
        conn.close()
    return render_template('register.html', error=error)

# Dashboard Page
@app.route('/dashboard')
def dashboard():
    user = {
        "name": "Siddharth",
    }
    return render_template('dashboard.html', user=user)

# Logout
@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/receive', methods=['GET', 'POST'])
def receive():
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if request.method == 'POST':
        category = request.form['category']
        item_name = request.form['item_name']
        head = request.form['head']
        ledger_page_no = request.form['ledger_page_no']
        available_stock = request.form['available_stock']
        qty = int(request.form['qty'])  # Ensure qty is an integer
        price_unit = request.form['price_unit']
        remarks = request.form['remarks']
        date = request.form['date']

        # Insert into received_items
        cur.execute("""INSERT INTO received_items 
            (category, item_name, head, ledger_page_no, available_stock, qty, price_unit, remarks, date)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (category, item_name, head, ledger_page_no, available_stock, qty, price_unit, remarks, date))

        # Calculate previous balance for this item
        cur.execute("SELECT COALESCE(SUM(receive_qty) - SUM(issue_qty), 0) as balance FROM ledger_entries WHERE item_name = %s", (item_name,))
        prev_balance = cur.fetchone()['balance'] or 0

        # Insert into ledger_entries
        new_balance = prev_balance + qty
        cur.execute("""INSERT INTO ledger_entries 
            (item_name, date, type, receive_from, issue_to, prev_bal, receive_qty, issue_qty, balance, remark)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (item_name, date, 'Receive', head, None, prev_balance, qty, 0, new_balance, remarks))

        conn.commit()
        conn.close()
        return redirect(url_for('receive'))

    cur.execute("SELECT * FROM received_items ORDER BY id DESC")
    entries = cur.fetchall()
    entries = [dict(row) for row in entries]
    categories = Ledger.get_categories()
    items_by_category = {cat: Ledger.get_items_by_category(cat) for cat in categories}
    conn.close()
    return render_template(
        'receive.html',
        categories=categories,
        items_by_category=items_by_category,
        entries=entries
    )

@app.route('/issue', methods=['GET', 'POST'])
def issue():
    error = None
    message = None
    form_data = {}
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if request.method == 'POST':
        category = request.form['category']
        item_name = request.form['item_name']
        head = request.form['head']
        ledger_page_no = request.form['ledger_page_no']
        available_stock = request.form['available_stock']
        qty = int(request.form['qty'])
        issued_to = request.form['issued_to']
        date = request.form['date']
        remarks = request.form['remarks']

        # Check available stock
        cur.execute("SELECT SUM(qty) as total_received FROM received_items WHERE category=%s AND item_name=%s", (category, item_name))
        received = cur.fetchone()['total_received'] or 0
        cur.execute("SELECT SUM(qty) as total_issued FROM issued_items WHERE category=%s AND item_name=%s", (category, item_name))
        issued = cur.fetchone()['total_issued'] or 0
        available = received - issued

        if qty > available:
            error = f"Issue quantity ({qty}) cannot be greater than available stock ({available})."
            form_data = request.form.to_dict()
        else:
            # Insert into issued_items
            cur.execute(""" 
                INSERT INTO issued_items (category, item_name, head, ledger_page_no, available_stock, qty, issued_to, date, remarks)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (category, item_name, head, ledger_page_no, available_stock, qty, issued_to, date, remarks))

            # Calculate previous balance for this item
            cur.execute("SELECT COALESCE(SUM(receive_qty) - SUM(issue_qty), 0) as balance FROM ledger_entries WHERE item_name = %s", (item_name,))
            prev_balance = cur.fetchone()['balance'] or 0

            # Insert into ledger_entries
            new_balance = prev_balance - qty
            cur.execute("""INSERT INTO ledger_entries 
                (item_name, date, type, receive_from, issue_to, prev_bal, receive_qty, issue_qty, balance, remark)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (item_name, date, 'Issue', None, issued_to, prev_balance, 0, qty, new_balance, remarks))

            conn.commit()
            message = "Entry added successfully."
            form_data = {}

    cur.execute("SELECT * FROM issued_items ORDER BY id DESC")
    entries = cur.fetchall()
    entries = [dict(row) for row in entries]
    categories = Ledger.get_categories()
    items_by_category = {cat: Ledger.get_items_by_category(cat) for cat in categories}
    offices = HeadOfficeManager.get_all_offices()
    conn.close()
    return render_template(
        'issue.html',
        categories=categories,
        items_by_category=items_by_category,
        entries=entries,
        error=error,
        message=message,
        form_data=form_data,
        offices=offices
    )

class UserManager:
    @staticmethod
    def get_all_users():
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id, username, name, cisf_no, rank, mobile FROM users")
        users = cur.fetchall()
        conn.close()
        return [dict(row) for row in users]

    @staticmethod
    def add_user(username, password, name, cisf_no, rank, mobile):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO users (username, password, name, cisf_no, rank, mobile) VALUES (%s, %s, %s, %s, %s, %s)",
            (username, password, name, cisf_no, rank, mobile)
        )
        conn.commit()
        conn.close()

    @staticmethod
    def delete_user(user_id):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
        conn.commit()
        conn.close()

@app.route('/add_user', methods=['GET', 'POST'])
def add_user():
    if 'user' not in session:
        return redirect(url_for('login'))
    message = None
    error = None
    if request.method == 'POST':
        name = request.form.get('name')
        cisf_no = request.form.get('cisf_no')
        rank = request.form.get('rank')
        mobile = request.form.get('mobile')
        username = request.form.get('username')
        password = request.form.get('password')
        if not (name and cisf_no and rank and mobile and username and password):
            error = "All fields are required."
        else:
            try:
                UserManager.add_user(username, password, name, cisf_no, rank, mobile)
                message = "User added successfully."
            except Exception as e:
                error = f"Error: {str(e)}"
    return render_template('add_user.html', message=message, error=error)

@app.route('/user_list')
def user_list():
    if 'user' not in session:
        return redirect(url_for('login'))
    users = UserManager.get_all_users()
    error = request.args.get('error')
    message = request.args.get('message')
    return render_template('user_list.html', users=users, error=error, message=message)

@app.route('/ledger', methods=['GET', 'POST'])
def ledger():
    if 'user' not in session:
        return redirect(url_for('login'))
    message = None
    error = None
    if request.method == 'POST':
        category = request.form['category']
        item_name = request.form['item_name']
        head = request.form['head']
        ledger_page_no = request.form['ledger_page_no']
        opening_date = request.form['opening_date']
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT id FROM ledger WHERE category=%s AND item_name=%s AND head=%s",
            (category, item_name, head)
        )
        duplicate = cur.fetchone()
        if duplicate:
            error = "A ledger entry with the same Category, Description, and Head already exists."
            cur.execute("SELECT category_name FROM items_category ORDER BY category_name")
            categories = [row['category_name'] for row in cur.fetchall()]
            cur.execute("SELECT head FROM head_office WHERE head IS NOT NULL AND head != '' ORDER BY head")
            heads = [row['head'] for row in cur.fetchall()]
            items = Ledger.get_all()
            conn.close()
            return render_template(
                'ledger.html',
                items=items,
                categories=categories,
                heads=heads,
                message=message,
                error=error,
                now=datetime.now
            )
        Ledger.add(category, item_name, head, ledger_page_no, opening_date)
        message = "Ledger entry added successfully."
        conn.close()
    items = Ledger.get_all()
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT category_name FROM items_category ORDER BY category_name")
    categories = [row['category_name'] for row in cur.fetchall()]
    cur.execute("SELECT head FROM head_office WHERE head IS NOT NULL AND head != '' ORDER BY head")
    heads = [row['head'] for row in cur.fetchall()]
    conn.close()
    return render_template(
        'ledger.html',
        items=items,
        categories=categories,
        heads=heads,
        message=message,
        error=error,
        now=datetime.now
    )

class Ledger:
    @staticmethod
    def get_all():
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM ledger")
        items = cur.fetchall()
        conn.close()
        return [dict(row) for row in items]

    @staticmethod
    def add(category, item_name, head, ledger_page_no, opening_date):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO ledger (category, item_name, head, ledger_page_no, opening_date) VALUES (%s, %s, %s, %s, %s)",
            (category, item_name, head, ledger_page_no, opening_date)
        )
        conn.commit()
        conn.close()

    @staticmethod
    def get_categories():
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT DISTINCT category FROM ledger")
        categories = [row['category'] for row in cur.fetchall()]
        conn.close()
        return categories

    @staticmethod
    def get_items_by_category(category):
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT item_name FROM ledger WHERE category = %s", (category,))
        items = [row['item_name'] for row in cur.fetchall()]
        conn.close()
        return items

def create_ledger_table():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ledger (
            id SERIAL PRIMARY KEY,
            category TEXT NOT NULL,
            item_name TEXT NOT NULL,
            head TEXT,
            ledger_page_no TEXT,
            opening_date TEXT
        )
    """)
    conn.commit()
    conn.close()

def create_received_items_table():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS received_items (
            id SERIAL PRIMARY KEY,
            category TEXT NOT NULL,
            item_name TEXT NOT NULL,
            head TEXT,
            ledger_page_no TEXT,
            available_stock TEXT,
            qty INTEGER,
            price_unit TEXT,
            remarks TEXT,
            date TEXT
        )
    """)
    conn.commit()
    conn.close()

def recreate_issued_items_table():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS issued_items")
    cur.execute("""
        CREATE TABLE issued_items (
            id SERIAL PRIMARY KEY,
            category TEXT NOT NULL,
            item_name TEXT NOT NULL,
            head TEXT,
            ledger_page_no TEXT,
            available_stock INTEGER,
            qty INTEGER,
            issued_to TEXT,
            date TEXT,
            remarks TEXT
        )
    """)
    conn.commit()
    conn.close()

def create_head_office_table():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS head_office (
            id SERIAL PRIMARY KEY,
            head TEXT,
            office_name TEXT
        )
    """)
    conn.commit()
    conn.close()

def create_ledger_entries_table():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ledger_entries (
            id SERIAL PRIMARY KEY,
            item_name TEXT,
            date TEXT,
            type TEXT,
            receive_from TEXT,
            issue_to TEXT,
            prev_bal INTEGER,
            receive_qty INTEGER,
            issue_qty INTEGER,
            balance INTEGER,
            remark TEXT
        )
    """)
    conn.commit()
    conn.close()

def create_items_category_table():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS items_category (
            id SERIAL PRIMARY KEY,
            category_name TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

# Update user table creation to include role
def create_users_table():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            name TEXT,
            cisf_no TEXT,
            rank TEXT,
            mobile TEXT
        )
    """)
    # Add role column if it does not exist
    try:
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS role TEXT DEFAULT 'user';")
        conn.commit()
    except Exception as e:
        print("Error adding role column:", e)
    conn.close()

@app.route('/update_ledger/<int:id>', methods=['GET', 'POST'])
def update_ledger(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if request.method == 'POST':
        category = request.form['category']
        item_name = request.form['item_name']
        head = request.form['head']
        ledger_page_no = request.form['ledger_page_no']
        opening_date = request.form['opening_date']
        cur.execute("""
            UPDATE ledger SET category=%s, item_name=%s, head=%s, ledger_page_no=%s, opening_date=%s
            WHERE id=%s
        """, (category, item_name, head, ledger_page_no, opening_date, id))
        conn.commit()
        conn.close()
        return redirect(url_for('ledger'))
    cur.execute("SELECT * FROM ledger WHERE id=%s", (id,))
    item = cur.fetchone()
    cur.execute("SELECT category_name FROM items_category ORDER BY category_name")
    categories = [row['category_name'] for row in cur.fetchall()]
    cur.execute("SELECT head FROM head_office WHERE head IS NOT NULL AND head != '' ORDER BY head")
    heads = [row['head'] for row in cur.fetchall()]
    conn.close()
    return render_template('update_ledger.html', item=item, categories=categories, heads=heads, now=datetime.now)

@app.route('/delete_ledger/<int:id>', methods=['POST'])
def delete_ledger(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM ledger WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('ledger'))

@app.route('/update_receive/<int:id>', methods=['GET', 'POST'])
def update_receive(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if request.method == 'POST':
        category = request.form['category']
        item_name = request.form['item_name']
        head = request.form['head']
        ledger_page_no = request.form['ledger_page_no']
        available_stock = request.form['available_stock']
        qty = request.form['qty']
        price_unit = request.form['price_unit']
        remarks = request.form['remarks']
        date = request.form['date']
        cur.execute("""
            UPDATE received_items SET category=%s, item_name=%s, head=%s, ledger_page_no=%s, available_stock=%s, qty=%s, price_unit=%s, remarks=%s, date=%s
            WHERE id=%s
        """, (category, item_name, head, ledger_page_no, available_stock, qty, price_unit, remarks, date, id))
        conn.commit()
        conn.close()
        return redirect(url_for('receive'))
    cur.execute("SELECT * FROM received_items WHERE id=%s", (id,))
    entry = cur.fetchone()
    # Get all heads for dropdown
    cur.execute("SELECT head FROM head_office WHERE head IS NOT NULL AND head != '' ORDER BY head")
    heads = [row['head'] for row in cur.fetchall()]
    conn.close()
    return render_template('update_receive.html', entry=entry, heads=heads)

@app.route('/delete_receive/<int:id>', methods=['POST'])
def delete_receive(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM received_items WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('receive'))

@app.route('/update_issue/<int:id>', methods=['GET', 'POST'])
def update_issue(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if request.method == 'POST':
        category = request.form['category']
        item_name = request.form['item_name']
        qty = int(request.form['qty'])
        issued_to = request.form['issued_to']
        date = request.form['date']
        remarks = request.form['remarks']
        cur.execute("""
            UPDATE issued_items SET category=%s, item_name=%s, qty=%s, issued_to=%s, date=%s, remarks=%s
            WHERE id=%s
        """, (category, item_name, qty, issued_to, date, remarks, id))
        conn.commit()
        conn.close()
        return redirect(url_for('issue'))
    cur.execute("SELECT * FROM issued_items WHERE id=%s", (id,))
    entry = cur.fetchone()
    # Get all offices for dropdown
    offices = HeadOfficeManager.get_all_offices()
    conn.close()
    return render_template('update_issue.html', entry=entry, offices=offices)

@app.route('/delete_issue/<int:id>', methods=['POST'])
def delete_issue(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM issued_items WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('issue'))

@app.route('/get_ledger_info', methods=['POST'])
def get_ledger_info():
    category = request.form['category']
    item_name = request.form['item_name']
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT head, ledger_page_no FROM ledger WHERE category=%s AND item_name=%s", (category, item_name))
    row = cur.fetchone()
    cur.execute("SELECT SUM(qty) as total_received FROM received_items WHERE category=%s AND item_name=%s", (category, item_name))
    received = cur.fetchone()['total_received'] or 0
    cur.execute("SELECT SUM(qty) as total_issued FROM issued_items WHERE category=%s AND item_name=%s", (category, item_name))
    issued = cur.fetchone()['total_issued'] or 0
    available_stock = received - issued
    conn.close()
    if row:
        return jsonify({
            'head': row['head'],
            'ledger_page_no': row['ledger_page_no'],
            'available_stock': available_stock
        })
    else:
        return jsonify({'head': '', 'ledger_page_no': '', 'available_stock': ''})

@app.route('/manage_head_office', methods=['GET', 'POST'])
def manage_head_office():
    if 'user' not in session:
        return redirect(url_for('login'))
    message = None
    head_form_value = ''
    office_form_value = ''
    head_id = ''
    office_id = ''
    if request.method == 'POST':
        form_type = request.form.get('form_type')
        action = request.form.get('action')
        if form_type == 'head':
            head = request.form.get('head', '').strip()
            head_id = request.form.get('head_id', '')
            if action == 'save' and head:
                HeadOfficeManager.add(head, '')
                message = "Head added successfully."
            elif action == 'update' and head_id and head:
                HeadOfficeManager.update(head_id, head, '')
                message = "Head updated successfully."
            elif action == 'delete' and head_id:
                HeadOfficeManager.delete(head_id)
                message = "Head deleted successfully."
            head_form_value = head
        elif form_type == 'office':
            office_name = request.form.get('office_name', '').strip()
            office_id = request.form.get('office_id', '')
            if action == 'save' and office_name:
                HeadOfficeManager.add('', office_name)
                message = "Office added successfully."
            elif action == 'update' and office_id and office_name:
                HeadOfficeManager.update(office_id, '', office_name)
                message = "Office updated successfully."
            elif action == 'delete' and office_id:
                HeadOfficeManager.delete(office_id)
                message = "Office deleted successfully."
            office_form_value = office_name
        elif form_type == 'head_select':
            head_id = request.form.get('head_id')
            head_row = HeadOfficeManager.get_by_id(head_id)
            if head_row:
                head_form_value = head_row['head']
                head_id = head_row['id']
        elif form_type == 'office_select':
            office_id = request.form.get('office_id')
            office_row = HeadOfficeManager.get_by_id(office_id)
            if office_row:
                office_form_value = office_row['office_name']
                office_id = office_row['id']
    heads = HeadOfficeManager.get_all_heads()
    offices = HeadOfficeManager.get_all_offices()
    return render_template(
        'manage_head_office.html',
        heads=heads,
        offices=offices,
        message=message,
        head_form_value=head_form_value,
        office_form_value=office_form_value,
        head_id=head_id,
        office_id=office_id
    )

@app.route('/update_head_office/<int:id>', methods=['GET', 'POST'])
def update_head_office(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        head = request.form['head']
        office_name = request.form['office_name']
        # Show warning before update (for demonstration, you can adjust logic as needed)
        if request.form.get('confirm') == 'yes':
            HeadOfficeManager.update(id, head, office_name)
            conn.close()
            flash("Head/Office updated.", "success")
            return redirect(url_for('manage_head_office'))
        elif request.form.get('confirm') == 'no':
            conn.close()
            flash("Update cancelled.", "info")
            return redirect(url_for('manage_head_office'))
        # Show confirmation page
        entry = HeadOfficeManager.get_by_id(id)
        warning_msg = (
            "⚠️ Selected Head or Office se related sabhi data update ho sakte hain. "
            "Are you sure you want to proceed?"
        )
        conn.close()
        return render_template(
            'confirm_update_head_office.html',
            entry=entry,
            warning_msg=warning_msg,
            play_sound=True
        )
    cur.execute("SELECT * FROM head_office WHERE id=%s", (id,))
    entry = cur.fetchone()
    conn.close()
    return render_template('update_head_office.html', entry=entry)

@app.route('/delete_head_office/<int:id>', methods=['GET', 'POST'])
def delete_head_office(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    entry = HeadOfficeManager.get_by_id(id)
    if not entry:
        flash("Head/Office not found.", "error")
        return redirect(url_for('manage_head_office'))
    if request.method == 'POST':
        confirm = request.form.get('confirm')
        if confirm == 'yes':
            HeadOfficeManager.delete(id)
            flash("Head/Office and all related data deleted.", "success")
            return redirect(url_for('manage_head_office'))
        else:
            flash("Deletion cancelled.", "info")
            return redirect(url_for('manage_head_office'))
    warning_msg = (
        "⚠️ Selected Head or Office se related sabhi data delete ho jayenge. "
        "Are you sure you want to proceed?"
    )
    return render_template(
        'confirm_delete_head_office.html',
        entry=entry,
        warning_msg=warning_msg,
        play_sound=True
    )

@app.route('/report')
def report():
    if 'user' not in session:
        return redirect(url_for('login'))
    category = request.args.get('category', '')
    item = request.args.get('item', '')
    head = request.args.get('head', '')
    office = request.args.get('office', '')
    from_date = request.args.get('from_date', '2000-01-01')
    to_date = request.args.get('to_date', datetime.now().strftime('%Y-%m-%d'))
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = []
    params = []
    if category:
        where.append("category=%s")
        params.append(category)
    if item:
        where.append("item_name=%s")
        params.append(item)
    if head:
        where.append("head=%s")
        params.append(head)
    # Fix: Only filter by issued_to for issued_items, not for received_items
    where_received = list(where)
    params_received = list(params)
    where_issued = list(where)
    params_issued = list(params)
    if office:
        where_issued.append("issued_to=%s")
        params_issued.append(office)
    where_received.append("date BETWEEN %s AND %s")
    params_received.extend([from_date, to_date])
    where_issued.append("date BETWEEN %s AND %s")
    params_issued.extend([from_date, to_date])
    where_clause_received = " AND ".join(where_received) if where_received else "1=1"
    where_clause_issued = " AND ".join(where_issued) if where_issued else "1=1"
    query = f"""
        SELECT date, category, 'Receive' as type, item_name as item, remarks as description, head, ledger_page_no as lp_no,
               NULL as previous, qty as received, NULL as issued, NULL as office
        FROM received_items
        WHERE {where_clause_received}
        UNION ALL
        SELECT date, category, 'Issue' as type, item_name as item, remarks as description, head, ledger_page_no as lp_no,
               NULL as previous, NULL as received, qty as issued, issued_to as office
        FROM issued_items
        WHERE {where_clause_issued}
        ORDER BY date DESC
    """
    cur.execute(query, params_received + params_issued)
    transactions = [dict(row) for row in cur.fetchall()]
    cur.execute("SELECT DISTINCT category FROM ledger")
    categories = [row['category'] for row in cur.fetchall()]
    cur.execute("SELECT DISTINCT item_name FROM ledger")
    items = [row['item_name'] for row in cur.fetchall()]
    cur.execute("SELECT DISTINCT head FROM ledger")
    heads = [row['head'] for row in cur.fetchall()]
    cur.execute("SELECT DISTINCT issued_to FROM issued_items")
    offices = [row['issued_to'] for row in cur.fetchall()]
    items_by_category = {cat: Ledger.get_items_by_category(cat) for cat in categories}

    # Calculate totals from filtered transactions (fix: use only filtered transactions)
    total_received = sum(t.get('received', 0) or 0 for t in transactions if t.get('type') == 'Receive')
    total_issued = sum(t.get('issued', 0) or 0 for t in transactions if t.get('type') == 'Issue')
    total_balance = total_received - total_issued

    conn.close()
    return render_template(
        'report.html',
        transactions=transactions,
        categories=categories,
        items=items,
        items_by_category=items_by_category,
        heads=heads,
        offices=offices,
        total_received=total_received,
        total_issued=total_issued,
        total_balance=total_balance,
        now=datetime.now
    )

class HeadOfficeManager:
    @staticmethod
    def get_all():
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM head_office")
        rows = cur.fetchall()
        conn.close()
        return [dict(row) for row in rows]

    @staticmethod
    def get_all_heads():
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id, head FROM head_office WHERE head != ''")
        rows = cur.fetchall()
        conn.close()
        return [dict(row) for row in rows]

    @staticmethod
    def get_all_offices():
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id, office_name FROM head_office WHERE office_name != ''")
        rows = cur.fetchall()
        conn.close()
        return [dict(row) for row in rows]

    @staticmethod
    def add(head, office_name):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO head_office (head, office_name) VALUES (%s, %s)", (head, office_name))
        conn.commit()
        conn.close()

    @staticmethod
    def update(id, head, office_name):
        conn = get_db_connection()
        cur = conn.cursor()
        if head:
            cur.execute("UPDATE head_office SET head=%s WHERE id=%s", (head, id))
        elif office_name:
            cur.execute("UPDATE head_office SET office_name=%s WHERE id=%s", (office_name, id))
        conn.commit()
        conn.close()

    @staticmethod
    def delete(id):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM head_office WHERE id=%s", (id,))
        conn.commit()
        conn.close()

    @staticmethod
    def get_by_id(id):
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM head_office WHERE id=%s", (id,))
        row = cur.fetchone()
        conn.close()
        return dict(row) if row else None

@app.route('/export_excel')
def export_excel():
    if 'user' not in session:
        return redirect(url_for('login'))
    category = request.args.get('category', '')
    item = request.args.get('item', '')
    head = request.args.get('head', '')
    office = request.args.get('office', '')
    from_date = request.args.get('from_date', '2000-01-01')
    to_date = request.args.get('to_date', datetime.now().strftime('%Y-%m-%d'))
    conn = get_db_connection()
    cur = conn.cursor()
    where = []
    params = []
    if category:
        where.append("category=%s")
        params.append(category)
    if item:
        where.append("item_name=%s")
        params.append(item)
    if head:
        where.append("head=%s")
        params.append(head)
    if office:
        where.append("issued_to=%s")
        params.append(office)
    where.append("date BETWEEN %s AND %s")
    params.extend([from_date, to_date])
    where_clause = " AND ".join(where) if where else "1=1"
    query = f"""
        SELECT date, 'Receive' as type, category, item_name as name, head, ledger_page_no as lp_no,
               NULL as previous, qty as received, NULL as issued, NULL as office
        FROM received_items
        WHERE {where_clause}
        UNION ALL
        SELECT date, 'Issue' as type, category, item_name as name, head, ledger_page_no as lp_no,
               NULL as previous, NULL as received, qty as issued, issued_to as office
        FROM issued_items
        WHERE {where_clause}
        ORDER BY date DESC
    """
    cur.execute(query, params * 2)
    rows = cur.fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaction Report"
    headers = ["Date", "Type", "Category", "Name", "Head", "L/P No.", "Previous", "Received", "Issue", "Office"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row[0], row[1], row[2], row[3], row[4], row[5],
            row[6] if row[6] is not None else '',
            row[7] if row[7] is not None else '',
            row[8] if row[8] is not None else '',
            row[9] if row[9] is not None else ''
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="transaction_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/export_ledger_excel')
def export_ledger_excel():
    if 'user' not in session:
        return redirect(url_for('login'))
    item = request.args.get('item', '')
    if not item:
        return "No item selected", 400
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM ledger_entries WHERE item_name = %s ORDER BY TO_DATE(date, 'YYYY-MM-DD')", (item,))
    rows = cur.fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"
    headers = ["S.No", "Date", "Type", "Receive From", "Issue To", "Prev Bal", "Receive Qty", "Issue Qty", "Balance", "Remark"]
    ws.append(headers)
    for idx, row in enumerate(rows, 1):
        ws.append([
            idx, row['date'], row['type'], row['receive_from'] or '-', row['issue_to'] or '-',
            row['prev_bal'] or 0, row['receive_qty'] or 0, row['issue_qty'] or 0, row['balance'] or 0, row['remark'] or '-'
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"ledger_{item}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/export_pdf')
def export_pdf():
    return "PDF export coming soon!"

def get_all_items():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT DISTINCT item_name FROM ledger")
    items = [row['item_name'] for row in cur.fetchall()]
    conn.close()
    return items

@app.route('/get_items_by_category', methods=['POST'])
def get_items_by_category():
    category = request.form['category']
    items = Ledger.get_items_by_category(category)
    return jsonify({'items': items})

@app.route('/print_ledger')
def print_ledger():
    category = request.args.get('category')
    item = request.args.get('item')
    item_info = None
    transactions = []
    # Fetch categories, items, and items_by_category for treeview and dropdowns
    categories = Ledger.get_categories()
    items_by_category = {cat: Ledger.get_items_by_category(cat) for cat in categories}
    items = []
    selected_item = item if item else ''
    if category:
        items = items_by_category.get(category, [])
    else:
        # If no category selected, flatten all items
        for item_list in items_by_category.values():
            items.extend(item_list)
    if category and item:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        # Ledger info
        cur.execute("SELECT * FROM ledger WHERE category=%s AND item_name=%s", (category, item))
        item_info = cur.fetchone()
        # Transactions
        cur.execute("SELECT * FROM ledger_entries WHERE item_name=%s ORDER BY TO_DATE(date, 'YYYY-MM-DD')", (item,))
        transactions = cur.fetchall()
        conn.close()
    return render_template(
        'print_ledger.html',
        item_info=item_info,
        transactions=transactions,
        categories=categories,
        items=items,
        items_by_category=items_by_category,
        selected_item=selected_item
    )

@app.route('/print_ledger_print')
def print_ledger_print():
    category = request.args.get('category')
    item = request.args.get('item')
    item_info = None
    transactions = []
    if category and item:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM ledger WHERE category=%s AND item_name=%s", (category, item))
        item_info = cur.fetchone()
        cur.execute("SELECT * FROM ledger_entries WHERE item_name=%s ORDER BY TO_DATE(date, 'YYYY-MM-DD')", (item,))
        transactions = cur.fetchall()
        conn.close()
    return render_template(
        'print_ledger_print.html',
        item_info=item_info,
        transactions=transactions,
        entries=transactions  # <-- add this line
    )

@app.route('/export_ledger_pdf')
def export_ledger_pdf():
    if 'user' not in session:
        return redirect(url_for('login'))
    item = request.args.get('item', '')
    if not item:
        return "No item selected", 400
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM ledger WHERE item_name = %s", (item,))
    item_info = cur.fetchone()
    cur.execute("SELECT * FROM ledger_entries WHERE item_name = %s ORDER BY TO_DATE(date, 'YYYY-MM-DD')", (item,))
    rows = cur.fetchall()
    conn.close()
    rendered = render_template(
        'print_ledger_pdf.html',
        item_info=item_info,
        transactions=rows
    )
    pdf = pdfkit.from_string(rendered, False)
    return send_file(
        io.BytesIO(pdf),
        as_attachment=True,
        download_name=f"ledger_{item}.pdf",
        mimetype="application/pdf"
    )

@app.route('/items_category', methods=['GET', 'POST'])
def items_category():
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    message = ""
    if request.method == 'POST':
        category_name = request.form['category_name']
        cur.execute("INSERT INTO items_category (category_name) VALUES (%s)", (category_name,))
        conn.commit()
        message = "Category added successfully."
    cur.execute("SELECT * FROM items_category ORDER BY id DESC")
    categories = cur.fetchall()
    conn.close()
    return render_template('items_category.html', categories=categories, message=message)

@app.route('/update_category/<int:id>', methods=['GET', 'POST'])
def update_category(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if request.method == 'POST':
        category_name = request.form['category_name'].strip()
        if category_name:
            cur.execute("UPDATE items_category SET category_name=%s WHERE id=%s", (category_name, id))
            conn.commit()
            conn.close()
            return redirect(url_for('items_category'))
    cur.execute("SELECT * FROM items_category WHERE id=%s", (id,))
    category = cur.fetchone()
    conn.close()
    return render_template('update_items_category.html', category=category)

@app.route('/delete_category/<int:id>', methods=['POST'])
def delete_category(id):
    if 'user' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM items_category WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('items_category'))

class RenewalVoucher:
    @staticmethod
    def get_all(office=None):
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if office:
            cur.execute("SELECT * FROM renewal_voucher WHERE office=%s ORDER BY id DESC", (office,))
        else:
            cur.execute("SELECT * FROM renewal_voucher ORDER BY id DESC")
        vouchers = cur.fetchall()
        conn.close()
        print("Fetched vouchers:", vouchers)  # Debug: See what is fetched
        return [dict(row) for row in vouchers]

    @staticmethod
    def add(item_name, quantity, date, remarks, head=None, lp_no=None, office=None):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO renewal_voucher (item_name, quantity, date, remarks, head, lp_no, office) VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (item_name, quantity, date, remarks, head, lp_no, office)
        )
        conn.commit()
        conn.close()

    @staticmethod
    def update(voucher_id, item_name, quantity, date, remarks, head=None, lp_no=None, office=None):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "UPDATE renewal_voucher SET item_name=%s, quantity=%s, date=%s, remarks=%s, head=%s, lp_no=%s, office=%s WHERE id=%s",
            (item_name, quantity, date, remarks, head, lp_no, office, voucher_id)
        )
        conn.commit()
        conn.close()

    @staticmethod
    def delete(voucher_id):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM renewal_voucher WHERE id=%s", (voucher_id,))
        conn.commit()
        conn.close()

    @staticmethod
    def get_by_id(voucher_id):
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM renewal_voucher WHERE id=%s", (voucher_id,))
        row = cur.fetchone()
        conn.close()
        return dict(row) if row else None

def create_renewal_voucher_table():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS renewal_voucher (
            id SERIAL PRIMARY KEY,
            item_name TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            date TEXT NOT NULL,
            remarks TEXT,
            head TEXT,
            lp_no TEXT,
            office TEXT
        )
    """)
    # Ensure 'office' column exists (for upgrades)
    cur.execute("ALTER TABLE renewal_voucher ADD COLUMN IF NOT EXISTS office TEXT;")
    conn.commit()
    conn.close()

@app.route('/renewal_voucher')
def renewal_voucher():
    office = request.args.get('office', '')
    offices = HeadOfficeManager.get_all_offices()
    vouchers = RenewalVoucher.get_all(office if office else None)
    return render_template('renewal_voucher.html', vouchers=vouchers, offices=offices, selected_office=office)

@app.route('/export_renewal_voucher_excel')
def export_renewal_voucher_excel():
    office = request.args.get('office', '')
    vouchers = RenewalVoucher.get_all(office if office else None)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Renewal Voucher"
    headers = ["SL No.", "Item Name & Description", "Head", "L/P No.", "Qty Issue", "Issue Date", "Remark", "Office"]
    ws.append(headers)
    for idx, v in enumerate(vouchers, 1):
        ws.append([
            idx, v['item_name'], v.get('head', ''), v.get('lp_no', ''), v['quantity'], v['date'], v.get('remarks', ''), v.get('office', '')
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="renewal_voucher.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/export_renewal_voucher_pdf')
def export_renewal_voucher_pdf():
    office = request.args.get('office', '')
    vouchers = RenewalVoucher.get_all(office if office else None)
    rendered = render_template('renewal_voucher_pdf.html', vouchers=vouchers, office=office)
    options = {
        'enable-local-file-access': None,
        'load-error-handling': 'ignore',
        'load-media-error-handling': 'ignore'
    }
    pdf = pdfkit.from_string(rendered, False, configuration=pdfkit_config, options=options)
    return send_file(
        io.BytesIO(pdf),
        as_attachment=True,
        download_name="renewal_voucher.pdf",
        mimetype="application/pdf"
    )

@app.route('/add_renewal_voucher', methods=['GET', 'POST'])
def add_renewal_voucher():
    message = None
    error = None
    offices = HeadOfficeManager.get_all_offices()
    if request.method == 'POST':
        item_name = request.form.get('item_name')
        quantity = request.form.get('quantity')
        date = request.form.get('date')
        remarks = request.form.get('remarks')
        head = request.form.get('head')
        lp_no = request.form.get('lp_no')
        office = request.form.get('office')
        if not (item_name and quantity and date and office):
            error = "Item name, quantity, date, and office are required."
        else:
            try:
                RenewalVoucher.add(item_name, quantity, date, remarks, head, lp_no, office)
                message = "Renewal voucher added successfully."
            except Exception as e:
                error = f"Error: {str(e)}"
    return render_template('add_renewal_voucher.html', message=message, error=error, offices=offices)

if __name__ == '__main__':
    create_users_table()
    create_ledger_table()
    create_received_items_table()
    create_items_category_table()
    create_head_office_table()
    create_ledger_entries_table()
    create_renewal_voucher_table()
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port)